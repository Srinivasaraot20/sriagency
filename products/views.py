from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from categories.models import Category
from .models import Product, ProductImage
from dashboard.views import role_required
import csv
import io
import os
from openpyxl import Workbook, load_workbook

@login_required
@role_required(['super_admin', 'admin', 'store_manager'])
def product_list(request):
    products = Product.objects.all().select_related('category')
    categories = Category.objects.all()

    # Search & Filter
    query = request.GET.get('q')
    category_id = request.GET.get('category')
    status = request.GET.get('status')
    featured = request.GET.get('featured')
    sort_by = request.GET.get('sort', '-created_at')

    if query:
        products = products.filter(
            Q(name__icontains=query) |
            Q(brand__icontains=query) |
            Q(sku__icontains=query)
        )
    if category_id:
        products = products.filter(category_id=category_id)
    if status == 'active':
        products = products.filter(is_active=True)
    elif status == 'inactive':
        products = products.filter(is_active=False)
    if featured == 'yes':
        products = products.filter(is_featured=True)

    # Sort
    products = products.order_by(sort_by)

    # Bulk actions
    if request.method == 'POST':
        action = request.POST.get('action')
        product_ids = request.POST.getlist('selected_products')
        if product_ids:
            selected = Product.objects.filter(id__in=product_ids)
            if action == 'delete':
                selected.delete()
                messages.success(request, f"Successfully deleted {len(product_ids)} products.")
            elif action == 'activate':
                selected.update(is_active=True)
                messages.success(request, f"Successfully activated {len(product_ids)} products.")
            elif action == 'deactivate':
                selected.update(is_active=False)
                messages.success(request, f"Successfully deactivated {len(product_ids)} products.")
            return redirect('product_list')

    context = {
        'products': products,
        'categories': categories,
        'query': query,
        'category_id': category_id,
        'status': status,
        'featured': featured,
        'sort_by': sort_by,
    }
    return render(request, 'dashboard/products/list.html', context)

@login_required
@role_required(['super_admin', 'admin', 'store_manager'])
def product_add(request):
    categories = Category.objects.all()
    if request.method == 'POST':
        try:
            category_id = request.POST.get('category')
            category = get_object_or_404(Category, id=category_id)
            
            product = Product.objects.create(
                name=request.POST.get('name'),
                category=category,
                brand=request.POST.get('brand'),
                product_type=request.POST.get('product_type'),
                suitable_crops=request.POST.get('suitable_crops'),
                short_description=request.POST.get('short_description'),
                full_description=request.POST.get('full_description'),
                features=request.POST.get('features'),
                benefits=request.POST.get('benefits'),
                usage_instructions=request.POST.get('usage_instructions'),
                mrp=request.POST.get('mrp'),
                selling_price=request.POST.get('selling_price'),
                discount_price=request.POST.get('discount_price') or None,
                sku=request.POST.get('sku'),
                stock_quantity=request.POST.get('stock_quantity') or 0,
                availability=request.POST.get('availability') == 'on',
                is_featured=request.POST.get('is_featured') == 'on',
                is_active=request.POST.get('is_active') == 'on',
                main_image=request.FILES.get('main_image'),
                image=request.FILES.get('main_image'),
                thumbnail=request.FILES.get('thumbnail'),
                meta_title=request.POST.get('meta_title'),
                meta_description=request.POST.get('meta_description'),
                seo_keywords=request.POST.get('seo_keywords')
            )
            
            # Handle multiple gallery images
            gallery_images = request.FILES.getlist('gallery_images')
            for img in gallery_images:
                ProductImage.objects.create(product=product, image=img)
                
            messages.success(request, f"Product '{product.name}' created successfully.")
            return redirect('product_list')
        except Exception as e:
            messages.error(request, f"Error creating product: {e}")
            
    return render(request, 'dashboard/products/form.html', {'categories': categories})

@login_required
@role_required(['super_admin', 'admin', 'store_manager'])
def product_edit(request, pk):
    product = get_object_or_404(Product, pk=pk)
    categories = Category.objects.all()
    
    if request.method == 'POST':
        try:
            category_id = request.POST.get('category')
            product.category = get_object_or_404(Category, id=category_id)
            product.name = request.POST.get('name')
            product.brand = request.POST.get('brand')
            product.product_type = request.POST.get('product_type')
            product.suitable_crops = request.POST.get('suitable_crops')
            product.short_description = request.POST.get('short_description')
            product.full_description = request.POST.get('full_description')
            product.features = request.POST.get('features')
            product.benefits = request.POST.get('benefits')
            product.usage_instructions = request.POST.get('usage_instructions')
            product.mrp = request.POST.get('mrp')
            product.selling_price = request.POST.get('selling_price')
            product.discount_price = request.POST.get('discount_price') or None
            product.sku = request.POST.get('sku')
            product.stock_quantity = request.POST.get('stock_quantity') or 0
            product.availability = request.POST.get('availability') == 'on'
            product.is_featured = request.POST.get('is_featured') == 'on'
            product.is_active = request.POST.get('is_active') == 'on'
            
            if request.POST.get('remove_main_image') == 'on':
                if product.main_image and os.path.exists(product.main_image.path):
                    try:
                        os.remove(product.main_image.path)
                    except Exception:
                        pass
                product.main_image = None
                product.image = None

            if request.POST.get('remove_thumbnail') == 'on':
                if product.thumbnail and os.path.exists(product.thumbnail.path):
                    try:
                        os.remove(product.thumbnail.path)
                    except Exception:
                        pass
                product.thumbnail = None

            if request.FILES.get('main_image'):
                product.main_image = request.FILES.get('main_image')
                product.image = request.FILES.get('main_image')

            if request.FILES.get('thumbnail'):
                product.thumbnail = request.FILES.get('thumbnail')
                
            product.meta_title = request.POST.get('meta_title')
            product.meta_description = request.POST.get('meta_description')
            product.seo_keywords = request.POST.get('seo_keywords')
            product.save()
            
            # Add new gallery images
            gallery_images = request.FILES.getlist('gallery_images')
            for img in gallery_images:
                ProductImage.objects.create(product=product, image=img)
                
            messages.success(request, f"Product '{product.name}' updated successfully.")
            return redirect('product_list')
        except Exception as e:
            messages.error(request, f"Error updating product: {e}")
            
    return render(request, 'dashboard/products/form.html', {
        'product': product,
        'categories': categories,
        'gallery': product.gallery.all()
    })

@login_required
@role_required(['super_admin', 'admin', 'store_manager'])
def product_delete(request, pk):
    product = get_object_or_404(Product, pk=pk)
    name = product.name
    product.delete()
    messages.success(request, f"Product '{name}' deleted successfully.")
    return redirect('product_list')

@login_required
@role_required(['super_admin', 'admin', 'store_manager'])
def product_delete_image(request, img_id):
    img = get_object_or_404(ProductImage, id=img_id)
    product_id = img.product.id
    img.delete()
    messages.success(request, "Gallery image removed.")
    return redirect('product_edit', pk=product_id)

@login_required
@role_required(['super_admin', 'admin', 'store_manager'])
def export_products(request, format_type):
    products = Product.objects.all().select_related('category')
    
    if format_type == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="products_export.csv"'
        writer = csv.writer(response)
        writer.writerow(['SKU', 'Name', 'Category', 'Brand', 'MRP', 'Selling Price', 'Stock Quantity', 'Is Active', 'Is Featured'])
        
        for p in products:
            writer.writerow([p.sku, p.name, p.category.name, p.brand, p.mrp, p.selling_price, p.stock_quantity, p.is_active, p.is_featured])
        return response
        
    elif format_type == 'excel':
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="products_export.xlsx"'
        wb = Workbook()
        ws = wb.active
        ws.title = "Products"
        
        # Headers
        ws.append(['SKU', 'Name', 'Category', 'Brand', 'MRP', 'Selling Price', 'Stock Quantity', 'Is Active', 'Is Featured'])
        
        for p in products:
            ws.append([p.sku, p.name, p.category.name, p.brand, float(p.mrp), float(p.selling_price), p.stock_quantity, p.is_active, p.is_featured])
            
        wb.save(response)
        return response
        
    return redirect('product_list')

@login_required
@role_required(['super_admin', 'admin', 'store_manager'])
def import_products(request):
    if request.method == 'POST' and request.FILES.get('import_file'):
        import_file = request.FILES['import_file']
        filename = import_file.name
        
        try:
            if filename.endswith('.csv'):
                data_set = import_file.read().decode('UTF-8')
                io_string = io.StringIO(data_set)
                next(io_string) # Skip header
                reader = csv.reader(io_string, delimiter=',', quotechar='"')
                success_count = 0
                for row in reader:
                    if len(row) < 7:
                        continue
                    sku, name, cat_name, brand, mrp, selling_price, stock_qty = row[0], row[1], row[2], row[3], row[4], row[5], row[6]
                    category, _ = Category.objects.get_or_create(name=cat_name)
                    
                    Product.objects.update_or_create(
                        sku=sku,
                        defaults={
                            'name': name,
                            'category': category,
                            'brand': brand,
                            'mrp': mrp,
                            'selling_price': selling_price,
                            'stock_quantity': stock_qty,
                            'is_active': True
                        }
                    )
                    success_count += 1
                messages.success(request, f"Successfully imported {success_count} products from CSV.")
                
            elif filename.endswith('.xlsx'):
                wb = load_workbook(import_file)
                ws = wb.active
                success_count = 0
                # Iterate rows skipping header
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or len(row) < 7:
                        continue
                    sku, name, cat_name, brand, mrp, selling_price, stock_qty = row[0], row[1], row[2], row[3], row[4], row[5], row[6]
                    if not sku or not name:
                        continue
                    category, _ = Category.objects.get_or_create(name=str(cat_name))
                    
                    Product.objects.update_or_create(
                        sku=str(sku),
                        defaults={
                            'name': str(name),
                            'category': category,
                            'brand': str(brand) if brand else '',
                            'mrp': mrp,
                            'selling_price': selling_price,
                            'stock_quantity': int(stock_qty) if stock_qty is not None else 0,
                            'is_active': True
                        }
                    )
                    success_count += 1
                messages.success(request, f"Successfully imported {success_count} products from Excel.")
            else:
                messages.error(request, "Unsupported file format. Please upload CSV or XLSX.")
        except Exception as e:
            messages.error(request, f"Failed to import: {e}")
            
    return redirect('product_list')
